"""
Modelo de dados central do Gerador de Cards de Formatura PROGRAD/UFSM.
"""
import csv
import io
import os
from dataclasses import dataclass, field
from typing import List, Union


@dataclass
class Formando:
    nome: str = ""
    curso: str = ""


@dataclass
class MembroMesa:
    nome: str = ""
    cargo: str = ""


@dataclass
class Homenageado:
    nome: str = ""
    titulo: str = ""


@dataclass
class DadosCerimonia:
    texto_fixo: str = "Colações de Grau UFSM"
    ano: str = "2025"
    semestre: str = "2"
    formandos: List[Formando] = field(default_factory=list)
    mesa: List[MembroMesa] = field(default_factory=list)
    homenageados: List[Homenageado] = field(default_factory=list)

    @property
    def texto_evento(self) -> str:
        return f"{self.texto_fixo} {self.ano}/{self.semestre}"

    def total_cards(self) -> int:
        return len(self.formandos) + len(self.mesa) + len(self.homenageados)

    def get_card(self, index: int):
        """Retorna o card na posição global (Formandos → Mesa → Homenageados)."""
        total_formandos = len(self.formandos)
        total_mesa = len(self.mesa)

        if index < total_formandos:
            f = self.formandos[index]
            return {
                "tipo": "formando",
                "numero": index + 1,
                "numero_str": f"{index + 1:02d}",
                "nome": f.nome,
                "subtitulo": f.curso,
                "rodape": self.texto_evento,
            }
        index -= total_formandos

        if index < total_mesa:
            m = self.mesa[index]
            return {
                "tipo": "mesa",
                "numero": index + 1,
                "numero_str": f"MESA - {index + 1}",
                "nome": m.nome,
                "subtitulo": m.cargo,
                "rodape": self.texto_evento,
            }
        index -= total_mesa

        h = self.homenageados[index]
        return {
            "tipo": "homenageado",
            "numero": index + 1,
            "numero_str": str(index + 1),
            "nome": h.nome,
            "subtitulo": h.titulo,
            "rodape": self.texto_evento,
        }


# ── Helpers CSV & Excel ───────────────────────────────────────────────────────

def _detect_delimiter(text: str) -> str:
    first_line = text.split('\n', 1)[0]
    if ';' in first_line:
        return ';'
    if '\t' in first_line:
        return '\t'
    return ','


def _normalizar_nome_curso(curso: str) -> str:
    if not curso:
        return ""
    c_upper = curso.strip().upper()
    if "FORMAÇÃO DE PROFESSORES" in c_upper or "PROGRAMA ESPECIAL" in c_upper or "PEG" in c_upper:
        return "PEG"
    return curso.strip()


def formatar_nome_curso(curso: str) -> str:
    """
    Aplica as quebras de linha específicas para os cursos compostos da UFSM:
    - ARTES CÊNICAS - DIREÇÃO TEATRAL -> ARTES CÊNICAS\nDIREÇÃO TEATRAL
    - ARTES CÊNICAS - INTERPRETAÇÃO TEATRAL -> ARTES CÊNICAS\nINTERPRETAÇÃO TEATRAL
    - COMUNICAÇÃO SOCIAL - JORNALISMO -> COMUNICAÇÃO SOCIAL\nJORNALISMO
    - COMUNICAÇÃO SOCIAL - PRODUÇÃO EDITORIAL -> COMUNICAÇÃO SOCIAL\nPRODUÇÃO EDITORIAL
    - COMUNICAÇÃO SOCIAL - PUBLICIDADE E PROPAGANDA -> COMUNICAÇÃO SOCIAL\nPUBLICIDADE E PROPAGANDA
    - COMUNICAÇÃO SOCIAL - RELAÇÕES PÚBLICAS -> COMUNICAÇÃO SOCIAL\nRELAÇÕES PÚBLICAS
    """
    if not curso:
        return ""

    c_norm = _normalizar_nome_curso(curso)

    import unicodedata
    s_clean = unicodedata.normalize('NFKD', c_norm).encode('ASCII', 'ignore').decode('utf-8').upper()

    # 1. Artes Cênicas
    if "ARTES" in s_clean:
        if "DIRECAO" in s_clean:
            return "ARTES CÊNICAS\nDIREÇÃO TEATRAL"
        elif "INTERPRETACAO" in s_clean or "TEATRAL" in s_clean or "CENICAS" in s_clean or "CANICAS" in s_clean:
            return "ARTES CÊNICAS\nINTERPRETAÇÃO TEATRAL"
        else:
            return "ARTES CÊNICAS"

    # 2. Comunicação Social
    if "COMUNICA" in s_clean:
        if "JORNALISMO" in s_clean:
            return "COMUNICAÇÃO SOCIAL\nJORNALISMO"
        elif "PRODUCAO" in s_clean or "EDITORIAL" in s_clean:
            return "COMUNICAÇÃO SOCIAL\nPRODUÇÃO EDITORIAL"
        elif "PUBLICIDADE" in s_clean or "PROPAGANDA" in s_clean:
            return "COMUNICAÇÃO SOCIAL\nPUBLICIDADE E PROPAGANDA"
        elif "RELACAO" in s_clean or "RELAGOES" in s_clean or "PUBLICAS" in s_clean:
            return "COMUNICAÇÃO SOCIAL\nRELAÇÕES PÚBLICAS"

    return c_norm


def parse_csv_formandos(text: str) -> List[Formando]:
    delim = _detect_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    result = []
    curso_atual = ""

    for row in reader:
        nome = ""
        curso = ""
        vals = [str(v).strip() for v in row.values() if v is not None and str(v).strip()]

        for k, v in row.items():
            if not k or not v:
                continue
            k_clean = str(k).strip().lower()
            if not nome and ("nome" in k_clean or "formando" in k_clean or "aluno" in k_clean):
                nome = str(v).strip()
            elif not curso and ("curso" in k_clean or "gradua" in k_clean):
                curso = str(v).strip()

        if not nome or not curso:
            if len(vals) == 1:
                # Banner de curso em 1 coluna
                v0 = vals[0]
                if not any(header_tag in v0.lower() for header_tag in ["ordem", "chamada", "localidade", "nome"]):
                    curso_atual = v0
                    continue
            elif len(vals) >= 2:
                if not nome:
                    nome = vals[1] if len(vals) >= 2 and "curso" not in vals[0].lower() else vals[0]
                if not curso:
                    curso = vals[0] if vals[0] != nome else (vals[1] if len(vals) > 1 else "")

        if not curso and curso_atual:
            curso = curso_atual

        curso = formatar_nome_curso(curso)

        if nome and not any(tag in nome.upper() for tag in ["NOME COMPLETO", "INSIRA POR", "LOCALIDADE"]):
            result.append(Formando(nome=nome, curso=curso))

    return result


def _obter_aba_excel(wb, palavras_chave: List[str]):
    """Procura a aba mais apropriada pelo nome (ex: FORMANDOS(AS)). Se não achar, usa a ativa."""
    for name in wb.sheetnames:
        name_clean = name.upper()
        if any(kw.upper() in name_clean for kw in palavras_chave):
            return wb[name]
    return wb.active


def parse_excel_formandos(file_path: str) -> List[Formando]:
    """
    Lê planilhas Excel (.xlsx, .xls) com suporte aos layouts da PROGRAD/UFSM:
    1. Seleção automática da aba 'FORMANDOS(AS)' ou similar.
    2. Layout 2 colunas convencionais (Nome, Curso).
    3. Layout com banners de seção de curso (ex: PEG, PEDAGOGIA DIURNO na col C) e formandos nas linhas seguintes.
    4. Layout PROGRAD (Col B = Curso por extenso, Col C = Nome do Formando, Col D = Cidade).
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = _obter_aba_excel(wb, ["FORMANDO", "FORMANDA", "ALUNO", "LISTA", "CARD"])
    result = []
    curso_atual = ""

    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue

        # Limpar células
        row_cells = [str(cell).strip() if cell is not None else "" for cell in row]

        # Ignorar linhas totalmente vazias
        non_empty = [c for c in row_cells if c]
        if not non_empty:
            continue

        row_str = " ".join(non_empty).upper()

        # Ignorar cabeçalhos de instrução da planilha (ex: INSIRA POR ORDEM ALFABÉTICA...)
        if any(h in row_str for h in ["INSIRA POR ORDEM", "PARA ANÚNCIO NA CHAMADA", "LOCALIDADE DE REFERÊNCIA"]):
            continue

        # Se for a linha de cabeçalho "CURSO | NOME COMPLETO | LOCALIDADE..."
        if "NOME COMPLETO" in row_str or ("CURSO" in row_str and "NOME" in row_str):
            continue

        # Detectar banner de curso destacado em 1 célula (ex: 'PEG', 'PEDAGOGIA DIURNO')
        # Geralmente em destaque na ColUNA C (índice 2) ou B (índice 1) onde outras colunas como D estão vazias
        if len(non_empty) == 1:
            val = non_empty[0]
            if val.upper() not in ["CURSO", "NOME COMPLETO", "NOME"]:
                curso_atual = val
            continue

        # Detectar layout oficial da planilha PROGRAD de Colações:
        # Coluna B (índice 1) = Curso por extenso
        # Coluna C (índice 2) = Nome Completo
        # Coluna D (índice 3) = Cidade / Referência
        col_b = row_cells[1] if len(row_cells) > 1 else ""
        col_c = row_cells[2] if len(row_cells) > 2 else ""
        col_d = row_cells[3] if len(row_cells) > 3 else ""

        # Verificar se ColUNA C é um banner de curso (ex: 'PEG', 'PEDAGOGIA DIURNO') e Col D está vazia
        if col_c and not col_b and not col_d:
            curso_atual = col_c
            continue
        if col_c and col_b == col_c and not col_d:
            curso_atual = col_c
            continue

        nome = ""
        curso = ""

        if col_c and (col_b or col_d or len(non_empty) >= 2):
            nome = col_c
            curso = col_b if col_b else curso_atual
        elif col_b and not col_c:
            nome = col_b
            curso = curso_atual
        elif len(non_empty) >= 2:
            nome = non_empty[1] if len(non_empty) > 1 else non_empty[0]
            curso = non_empty[0] if non_empty[0] != nome else (non_empty[1] if len(non_empty) > 1 else "")

        if not curso and curso_atual:
            curso = curso_atual

        curso = formatar_nome_curso(curso)

        if nome and nome.upper() not in ["NOME COMPLETO", "NOME", "CURSO"]:
            result.append(Formando(nome=nome, curso=curso))

    return result


def parse_arquivo_formandos(file_path: str, content: str = None) -> List[Formando]:
    """Ponto de entrada único para carregar Formandos a partir de arquivo .xlsx, .xls ou .csv."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".xlsx", ".xls"]:
        return parse_excel_formandos(file_path)
    else:
        if content is None:
            with open(file_path, encoding="utf-8-sig", errors="ignore") as f:
                content = f.read()
        return parse_csv_formandos(content)


def parse_excel_mesa(file_path: str) -> List[MembroMesa]:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = _obter_aba_excel(wb, ["MESA", "HONRA"])
    result = []

    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue
        row_cells = [str(cell).strip() if cell is not None else "" for cell in row]
        non_empty = [c for c in row_cells if c]
        if len(non_empty) < 2:
            continue

        c0_upper = non_empty[0].upper()
        if "NOME" in c0_upper or "CARGO" in c0_upper:
            continue

        nome = non_empty[0]
        cargo = non_empty[1]
        if nome:
            result.append(MembroMesa(nome=nome, cargo=cargo))

    return result


def parse_arquivo_mesa(file_path: str, content: str = None) -> List[MembroMesa]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".xlsx", ".xls"]:
        return parse_excel_mesa(file_path)
    else:
        if content is None:
            with open(file_path, encoding="utf-8-sig", errors="ignore") as f:
                content = f.read()
        return parse_csv_mesa(content)


def parse_excel_homenageados(file_path: str) -> List[Homenageado]:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = _obter_aba_excel(wb, ["HOMENAGEAD", "HOMENA"])
    result = []

    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue
        row_cells = [str(cell).strip() if cell is not None else "" for cell in row]
        non_empty = [c for c in row_cells if c]
        if len(non_empty) < 2:
            continue

        c0_upper = non_empty[0].upper()
        if "NOME" in c0_upper or "TÍTULO" in c0_upper or "TITULO" in c0_upper:
            continue

        nome = non_empty[0]
        titulo = non_empty[1]
        if nome:
            result.append(Homenageado(nome=nome, titulo=titulo))

    return result


def parse_arquivo_homenageados(file_path: str, content: str = None) -> List[Homenageado]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".xlsx", ".xls"]:
        return parse_excel_homenageados(file_path)
    else:
        if content is None:
            with open(file_path, encoding="utf-8-sig", errors="ignore") as f:
                content = f.read()
        return parse_csv_homenageados(content)


def parse_csv_mesa(text: str) -> List[MembroMesa]:
    delim = _detect_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    result = []
    for row in reader:
        nome = ""
        cargo = ""
        for k, v in row.items():
            if not k or not v:
                continue
            k_clean = str(k).strip().lower()
            if not nome and ("nome" in k_clean or "membro" in k_clean):
                nome = str(v).strip()
            elif not cargo and ("cargo" in k_clean or "fun" in k_clean or "tít" in k_clean or "tit" in k_clean):
                cargo = str(v).strip()

        if not nome or not cargo:
            vals = [str(v).strip() for v in row.values() if v is not None and str(v).strip()]
            if not nome and len(vals) >= 1:
                nome = vals[0]
            if not cargo and len(vals) >= 2:
                cargo = vals[1] if vals[0] == nome else vals[0]

        if nome:
            result.append(MembroMesa(nome=nome, cargo=cargo))
    return result


def parse_csv_homenageados(text: str) -> List[Homenageado]:
    delim = _detect_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    result = []
    for row in reader:
        nome = ""
        titulo = ""
        for k, v in row.items():
            if not k or not v:
                continue
            k_clean = str(k).strip().lower()
            if not nome and ("nome" in k_clean or "homenageado" in k_clean):
                nome = str(v).strip()
            elif not titulo and ("tit" in k_clean or "tít" in k_clean or "cargo" in k_clean or "turma" in k_clean):
                titulo = str(v).strip()

        if not nome or not titulo:
            vals = [str(v).strip() for v in row.values() if v is not None and str(v).strip()]
            if not nome and len(vals) >= 1:
                nome = vals[0]
            if not titulo and len(vals) >= 2:
                titulo = vals[1] if vals[0] == nome else vals[0]

        if nome:
            result.append(Homenageado(nome=nome, titulo=titulo))
    return result


def export_csv_formandos(formandos: List[Formando]) -> str:
    out = io.StringIO()
    w = csv.writer(out, delimiter=';')
    w.writerow(["Nome", "Curso"])
    for f in formandos:
        w.writerow([f.nome, f.curso])
    return out.getvalue()


def export_csv_mesa(mesa: List[MembroMesa]) -> str:
    out = io.StringIO()
    w = csv.writer(out, delimiter=';')
    w.writerow(["Nome", "Cargo"])
    for m in mesa:
        w.writerow([m.nome, m.cargo])
    return out.getvalue()


def export_csv_homenageados(homenageados: List[Homenageado]) -> str:
    out = io.StringIO()
    w = csv.writer(out, delimiter=';')
    w.writerow(["Nome", "Titulo"])
    for h in homenageados:
        w.writerow([h.nome, h.titulo])
    return out.getvalue()
